# GOAL
# a time tracker app
# user to select the task from a task list drop down (data source for task list -> excel)
# has buttons -> start, pause, end, reset
# stores the start time, end time in Excel on click of end button
# resets timer on reset button click
import customtkinter as ctk
import os
import pandas as pd
from PIL.ImageFile import ImageFile
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import datetime as dt
from enum import Enum
# for handling icons
from PIL import Image, ImageDraw
# for handling file paths
import sys
# to create a separate thread for sys tray icon
import threading
# to create a sys tray icon and to create menu items for sys tray icon right click
import pystray


class TimerStatus(Enum):
    # to track the timer status for the buttons to work correctly
    RUNNING = 1
    PAUSED = 2
    STOPPED = 3

class TaskTimer:
    def __init__(self) -> None:
        # set the window theme to 'dark' mode
        ctk.set_appearance_mode("dark")
        # initialize the main window
        self.app = ctk.CTk()
        self.app_title = "Time Keeper"
        self.app.title(self.app_title)
        # disable window resizing
        self.app.resizable(False, False)
        # to disable the toolbar and make it as a widget,
        # makes a window borderless and removes from taskbar
        self.app.overrideredirect(True)

        # -----------assets-----------
        # Excel file to store the task list, time
        self.excel_file = "Task_Timer.xlsx"
        # icon for the system tray
        self.app_icon = "app_icon.ico"
        # Sheet in the Excel file to store the task list
        self.excel_tasks_sheet = "Tasks"
        # name of the column storing tasks inside the Tasks sheet
        self.tasks_col_name = "Tasks"
        # Sheet in the Excel file to store the time for each task
        self.excel_time_sheet = "Time"
        # excel icon for excel_btn
        self.excel_btn_icon = "excel_btn_icon.png"

        # get the task list from the Excel if it exists, to populate task_list_menu combobox dropdown
        self.task_list = self._get_task_list()
        # task selected from the task_list_menu combobox
        self.current_task = ""

        # track the timer status - running, paused, stopped
        self.is_timer_running: TimerStatus = TimerStatus.STOPPED
        # to display timer text inside the timer_display Entry
        self.timer_text = ctk.StringVar()
        # to track the number of seconds elapsed and to use to set the text for timer_display Entry via timer_text
        # self.seconds_elapsed_ui = 0
        self.task_start_time = None # to store the start time of the task
        self.task_end_time = None # to store the end time of the task
        # track the start and end times of a task between pauses, if any
        self.segment_start = None
        self.segment_end = None
        # seconds to be accumulated excluding pause_seconds
        self.seconds_accumulated = 0.0

        # to manage placeholder text in the notes_entry field
        # when is_placeholder_active is True, show PH text in the notes_entry filed
        self.is_placeholder_active = True

        # declared it here as these are used by multiple methods
        self.task_list_menu = None
        self.status_label = None
        self.timer_display = None
        self.start_btn = None
        # Textbox for user to type in task notes
        self.notes_textbox = None

        # to store the after() ID and to handle .after() calls overlaps i.e., to be used in .after_cancel()
        self.status_update_queue = None
        self.timer_running_queue = None

        # to handle drag and reposition of the app window
        self.start_mouse_x_root = None
        self.start_mouse_y_root = None
        self.start_window_x_root = None
        self.start_window_y_root = None

        # to enable running and controlling the app from the system tray
        self.systray_icon = None
        # create a separate thread for sys tray icon run so that mainloop() does not block this
        self.systray_thread = threading.Thread(target=self._initialize_systray_icon, daemon=True)
        # start the sys tray thread
        self.systray_thread.start()

        # build the ui (widgets) of the app
        self._build_ui()

        # position app window in the bottom right corner of the screen
        self.position_window()
        # ensure the app is brought to the top on start
        self.app.attributes("-topmost", True)
        # self.app.attributes("-topmost", False)
        # retain the top position for 500 ms and release after that
        self.app.after(500, lambda: self.app.attributes('-topmost', False))

        self.app.mainloop()


    def _get_task_list(self):
        """
        Read tasks from the Excel file if it exists and is not empty
        Only includes tasks where Status='Active'
        :returns list: A list of tasks always starting with "<Add new task...>".
        Returns just ["<Add new task...>"] if the Excel file doesn't exist, is empty, or has invalid data.
        """

        tasks_list = []

        # 1. check if the file exists and is not empty
        if os.path.exists(self.excel_file) and os.path.getsize(self.excel_file) > 0:
            # to catch errors in reading the file
            try:
                # 2. read the Excel sheet into a DF
                tasks_df = pd.read_excel(self.excel_file, sheet_name=self.excel_tasks_sheet)
                # print(f"tasks df:\n{tasks_df}")

                # 3. check if DF is not empty (e.g., only headers)
                if not tasks_df.empty:
                    # 4. drop blanks in 'Tasks' column and filter out tasks with status != 'active'
                    active_tasks = tasks_df[
                        tasks_df[self.tasks_col_name].notna() &
                        (tasks_df["Status"].str.lower() == "active")
                    ]

                    # 5. convert the tasks to a list
                    tasks_list = active_tasks[self.tasks_col_name].to_list()
                    tasks_list.sort()

                else:
                    print(f"{self.excel_file} exists but contains no data.")
            except pd.errors.ParserError as e:
                print(f"Error reading the file to get the task list: {e}")
            except Exception as e:
                # catch any other unexpected errors
                print(f"An unexpected error occurred while getting the task list: {e}")
        else:
            print(f"{self.excel_file} doesn't exist or is empty.")

        # print(f"{tasks_list=}")
        return ["<Add new task...>"] + tasks_list


    def _list_menu_callback(self, choice):
        # instead of removing the existing task manually, if any,
        # user can select the "<Add new task...>" item to clear the field and set the focus to type the new task
        # space before < to ensure this stays at the top after sorting the list
        if choice == "<Add new task...>":
            self.task_list_menu.set("")
            # set the current_task to "" so that the previously set task is cleared, else we can still run the timer without typing or selecting a new task
            # Flow - select a task from the dropdown (task_list_menu combobox) -> current_task = selected_task due to else block
            # select "<Add new task...>" from the dropdown, task_list_menu will be set to "" and cursor will be blinking, but the current_task still has a value from previous selection
            # so, even without typing in a new task or selecting one from the dropdown, use can start the timer with start_btn as the check 'if self.current_task' in run_timer evaluates true
            # hence we have to clear the current_task whenever '"<Add new task...>"' is selected
            self.current_task = ""
            self.task_list_menu.focus()
        else:
            # set the current task value
            self.current_task = choice
            # to remove the blinking cursor in the combobox after item selection
            self.app.focus()
        # print("Selected Task:", choice)


    def _update_status_label(self, status: str, code: int):
        """
        Display status of a task for 3-seconds
        :param status: str
        :param code: int | 0 for success 1 for failure/error/warning
        """
        # check if a status update task is running and cancel it before starting a new task
        if self.status_update_queue is not None:
            self.app.after_cancel(self.status_update_queue)

        # show the task addition status
        if code == 1:
            # if there is an error/failure
            status = status + " :(" if status else status
            self.status_label.configure(text=status, text_color="#b54747")
        else:
            # if the task is successful
            status = status + " :)" if status else status
            self.status_label.configure(text=status, text_color="#009933")

        # schedule a new status update task to hide the status and store the ID
        # schedule task only if the status is not empty, else it will lead to infinite loop
        # Call 1 ("Added") → Call 2 ("") → Call 3 ("") → Call 4 ("") → ...
        if status:
            self.status_update_queue = self.app.after(3000,
                                                      lambda: self._update_status_label("", 1)
                                                      )


    def _append_data_to_excel(self, sheet_name, **kwargs) -> bool:
        """
        Appends new row of data to the specified sheet in the Excel file
        Creates the Excel file/new sheet if they don't exist
        Data is passed as keyword arguments and keywords become headers
        Example usage:
            _append_data_to_excel("Tasks", Task="Study Python", Status="Active", Added_On="2025-04-05 10:00")
            _append_data_to_excel("Time", Task="Study Python", Timestamp="2025-04-05 10:00", Notes="Great progress!")
        :param sheet_name: (str): Name of the sheet to append to
        :param kwargs: Each key becomes a column header, value becomes cell data
        :returns bool: True if successful, False otherwise
        """
        # using openpyxl for appending data as calculating last row (with openpyxl) is required for pandas
        try:
            # 1. check if the file exists or to be created
            if os.path.exists(self.excel_file):
                try:
                    # Attempt to load the workbook. This will fail for zero-byte or corrupted files
                    wb = load_workbook(self.excel_file)
                except (InvalidFileException, Exception) as e:
                    print(f"Error with existing file: {e}. Creating new file.")
                    wb = Workbook()
                    # remove default sheets created e.g., 'Sheet1'
                    if len(wb.sheetnames) > 0:
                        for s in wb.sheetnames:
                            wb.remove(wb[s])
            else:
                # file does not exist, so creating a new file
                wb = Workbook()
                # remove default sheets created e.g., 'Sheet1'
                if len(wb.sheetnames) > 0:
                    for s in wb.sheetnames:
                        wb.remove(wb[s])

            # 2. check if the sheet exists or to be created
            if sheet_name not in wb.sheetnames:
                sheet = wb.create_sheet(sheet_name)
                # set the headings for the sheet
                sheet.append(list(kwargs.keys()))
            else:
                # sheet exists in the Excel file
                sheet = wb[sheet_name]

            # 3. append the new data
            sheet.append(list(kwargs.values()))

            # 4. save and close the Excel file
            wb.save(self.excel_file)
            wb.close()
            return True
        except Exception as e:
            # This outer catch is for errors during sheet creation, appending, or saving
            print(f"Error on appending data to excel: {e}")
            return False


    def _add_task_on_enter(self, event):
        """
        To add a new task typed into the task_list_menu combobox to the task_list and Excel, on press of 'Enter' key
        """
        # get the text currently in the combobox entry
        # .strip() removes leading/trailing whitespace
        new_task = self.task_list_menu.get().strip()

        # if the new_task is not empty and does not exist in the task_list, add it to the task_list
        if new_task:
            # to preserve formats like 'ITR' 'GPS'
            new_task = new_task[0].upper() + new_task[1:]
            # for robust check on if the task exists
            task_list_lower = [task.lower() for task in self.task_list]
            if new_task.lower() not in task_list_lower:
                # add/append the task to excel and if that is successful, proceed further
                now = dt.datetime.now()
                write_status = self._append_data_to_excel(self.excel_tasks_sheet, Tasks=new_task, Status="Active", Added_On = f"{now:%d-%b-%Y T%I:%M %p}")

                # perform further steps
                if write_status:
                    self.task_list = self._get_task_list()
                    # to ensure "<Add new task...>" is at the top of the list
                    self.task_list[1:] = sorted(self.task_list[1:])
                    # update the combobox with the new task_list
                    self.task_list_menu.configure(values=self.task_list)
                    # set the value to new_task with spaces stripped and capitalized
                    self.task_list_menu.set(new_task)
                    # update the current task selection which will be used as validation for starting timer on click of start_btn in run_timer method
                    self.current_task = new_task
                    # show the task addition status
                    self._update_status_label("Added", 0)
                    # to remove focus (cursor) from the task_list_menu combobox
                    self.app.focus()
                else:
                    self._update_status_label("Error", 1)
            else:
                self._update_status_label("Exists", 1)
        else:
            self._update_status_label("Empty", 1)


    def _update_timer_display(self):
        """
        calculates the seconds elapsed and updates the UI
        """
        seconds_elapsed_ui = self.seconds_accumulated + (dt.datetime.now() - self.segment_start).total_seconds()
        seconds_elapsed_ui = int(seconds_elapsed_ui)

        hours_elapsed, remainder = divmod(seconds_elapsed_ui, 3600)
        # the remainder we get here is the seconds remaining
        minutes_elapsed, remainder = divmod(remainder, 60)
        # show the timer in the timer_display Entry via timer_text instance variable
        self.timer_text.set(f"{hours_elapsed:02}:{minutes_elapsed:02}:{remainder:02}")
        self._schedule_update_timer()


    def _schedule_update_timer(self):
        # recursively count seconds and update timer text as long as the timer is running
        # to ensure the first second is displayed after a second has actually elapsed
        if self.is_timer_running == TimerStatus.RUNNING:
            self.timer_running_queue = self.app.after(1000, self._update_timer_display)


    def _show_placeholder(self):
        """
        sets the placeholder text in the notes_entry field
        """
        self.notes_textbox.insert("1.0", "Add notes")
        self.notes_textbox.configure(text_color="#7a848d")
        # we set the status here instead of in _notes_focus_out() method so that when we call this method in _reset_timer(), even the status is also set
        self.is_placeholder_active = True


    def _notes_focus_in(self, event):
        """
        when notes_textbox is focused and if is_placeholder_active = True
        clear the placeholder text and change font color #7a848d -> #f2f2f2
        """
        print(f"_notes_focs_in: {event}")
        if self.is_placeholder_active:
            self.notes_textbox.delete("1.0", "end")
            self.notes_textbox.configure(text_color="#f2f2f2")
            self.is_placeholder_active = False


    def _notes_focus_out(self, event):
        """
        when notes_textbox loses focus, checks if there is a user entered text in it
        if there is no user entered text, then the placeholder text is shown
        """
        print(f"_notes_focs_out: {event}")
        if not self.notes_textbox.get("1.0", "end-1c").strip():
            # user has not entered any text or entered just spaces
            self._show_placeholder()
            # self.is_placeholder_active = True


    def _accumulator(self):
        """
        calculates the duration between segment_start and segment_end times and stores the duration in seconds_accumulated
        """
        self.seconds_accumulated += (self.segment_end - self.segment_start).total_seconds()


    def _run_timer(self):
        """
        Handles starting, pausing, resuming timer
        """
        # if self.start_btn.cget("text") == "▶":
        # check if a task is selected before starting the timer
        if self.current_task:
            # run timer if the timer is not running i.e., timer is paused or stopped
            if self.is_timer_running != TimerStatus.RUNNING:
                # we set the task_start_time before changing the is_timer_running status to catch the scenario where a timer that was paused is being resumed
                # when paused, is_timer_running = PAUSED, so != RUNNING
                # we then set is_timer_running = RUNNING and after that we again set task_start_time as this inner if evaluates true,
                # i.e., RUNNING != PAUSED which means a new task_start_time is created for an already running task
                # Set task_start_time only if the timer is beginning a new session (from STOPPED);
                # this preserves the original start time when resuming from a PAUSED state.
                # only set if completely new task
                if self.is_timer_running == TimerStatus.STOPPED:
                    self.task_start_time = dt.datetime.now()
                # set the start time of the segment (at the first run, i.e., starting when TimerStatus.STOPPED)
                self.segment_start = dt.datetime.now()
                self.is_timer_running = TimerStatus.RUNNING
                self.start_btn.configure(text="⏸")
                # to not select a new task while the timer is running
                self.task_list_menu.configure(state="disabled")
                self._update_timer_display()
                self._update_status_label("Start", 0)
                # print("Timer running")
            else:
                self.is_timer_running = TimerStatus.PAUSED
                # store the time when the task is paused
                self.segment_end = dt.datetime.now()
                # store the work seconds before till task is paused
                self._accumulator()
                self.start_btn.configure(text="▶")
                # as we are updating self.is_timer_running = TimerStatus.STOPPED, _update_timer() will stop as it runs only when the timer is running -> if self.is_timer_running == TimerStatus.RUNNING: ...
                # self._update_timer()
                self._update_status_label("Pause", 0)
                # without this, the previous scheduled call in _update_timer() keeps running/lingering even when paused and continue after resumed
                if self.timer_running_queue:
                    self.app.after_cancel(self.timer_running_queue)
                # print("Timer paused")
        else:
            # if no task is selected before starting the timer (hitting start_btn)
            self._update_status_label("Select", 1)
        self.app.focus()


    def _humanize_time(self, seconds) -> str:
        """
        Formats the total seconds elapsed (self.seconds_elapsed) into H:MM string format to be used for saving to the Excel
        Handles durations longer than 24 hours by accumulating hours
        Calculates timedelta = task_end_time - task_start_time and returns the time difference in a human-readable form - 2h:12m
        :return: str: 2h:12m
        """
        # this is incorrect as there may be pause time in between
        # difference = self.task_end_time - self.task_start_time
        # diff_seconds = difference.total_seconds()
        # seconds_elapsed represents only the time in seconds timer ran and not paused
        if seconds:
            hours, remainder = divmod(seconds, 3600)
            minutes, remainder = divmod(remainder, 60)

            return f"{hours:.0f}h:{minutes:02.0f}m:{remainder:02.0f}"
        else:
            return ""


    def _get_pause_duration(self) -> int:
        """
        checks if there are any pauses during the task and returns the total pause duration in seconds
        :return: int seconds
        """
        total_task_seconds = (self.task_end_time - self.task_start_time).total_seconds()
        pause_seconds = total_task_seconds - self.seconds_accumulated

        return pause_seconds

    # if there is an error in saving the data to the Excel file (e.g., file is opened and so permission is denied), we have to stop timer and show 'Error' status
    # when user clicks stop_btn again, we have to try saving to the Excel again (e.g., user closed the file now and hit stop_btn again)
    # since we set the running status = STOPPED in the else block also to stop the timer, next time when user clicks stop_btn, we will not try to save the data
    def _end_timer(self):
        """
        Stops the timer, saves the task log to Excel, and resets the timer state
        stop timer (_update_timer())
        reset the timer text (timer_text)
        change the timer running status (is_timer_running)
        change the symbol on the start button
        deselect task in task_list_menu
        """
        # if the timer is not stopped i.e., is_timer_status is running/paused
        if self.is_timer_running != TimerStatus.STOPPED:
            # to capture when the task has ended and to be logged to the Excel
            self.task_end_time = dt.datetime.now()
            # calculate the seconds accumulated if the is_timer_status is not paused
            # as the _accumulator is called only when paused
            if self.is_timer_running != TimerStatus.PAUSED:
                # as the timer is still running (i.e., not paused), the end time of the segment would be same as the end time of the task
                self.seconds_accumulated += (self.task_end_time - self.segment_start).total_seconds()

            # get the work duration in hh:mm format
            work_duration = self._humanize_time(self.seconds_accumulated)
            # get pause seconds only after final seconds_accumulated is arrived as pause_seconds is nothing but total_seconds - work_seconds (i.e., seconds_accumulated)
            pause_seconds = self._get_pause_duration()
            pause_duration = self._humanize_time(pause_seconds)

            # to avoid capturing placeholder text as notes
            if self.is_placeholder_active:
                task_notes = ""
            else:
                task_notes = self.notes_textbox.get("1.0", "end-1c")

            # write data to the Excel Date, Task, Duration, Notes, Start Time, End Time, Seconds
            # print(f"{self.task_start_time:%d-%b-%Y}, {self.current_task}, {self._humanize_time()}, {self.notes_entry.get('1.0', 'end-1c')},{self.task_start_time:%I:%M %p}, {self.task_end_time:%I:%M %p}, {self.seconds_elapsed}")
            write_status = self._append_data_to_excel(self.excel_time_sheet,
                                                      Date=f"{self.task_start_time:%d-%b-%Y}",
                                                      Task=self.current_task,
                                                      Work_Duration=work_duration,
                                                      Notes=task_notes,
                                                      Pause_Duration=pause_duration,
                                                      Start_Time=f"{self.task_start_time:%I:%M:%S %p}",
                                                      End_Time=f"{self.task_end_time:%I:%M:%S %p}",
                                                      Work_Seconds=self.seconds_accumulated,
                                                      Pause_Seconds=pause_seconds,
                                                      Total_Seconds= (self.task_end_time - self.task_start_time).total_seconds()
                                                      )
            # show status of saving the data to the Excel file
            if write_status:
                self._update_status_label("Saved", 0)
                self._reset_timer()
                self.current_task=""
                # we can't edit combobox when the state is disabled
                # the state is disabled when start_btn is clicked,
                # the state is set to normal inside the reset_timer,
                # so after that we can set the value to "", else this line will have no change
                # Gemini AI or qwen did not catch this
                self.task_list_menu.set("")
                print("Time log saved successfully.")
            else:
                # if failed to save the data to the Excel
                self._update_status_label("Error", 1)
                # if we set this to STOPPED, we can't attempt to retry saving to the Excel file
                self.is_timer_running = TimerStatus.PAUSED
                self.start_btn.configure(text="▶")
                if self.timer_running_queue:
                    self.app.after_cancel(self.timer_running_queue)
                print("Error saving the log!")
            # we reset the timer at the last so that seconds_elapsed is not set to 0 in the Excel

        # print("Timer stopped")


    def _reset_timer(self, status=""):
        """
        Resets the timer to its initial stopped state
        stop timer
        reset seconds to 0
        change timer running status, stop button symbol
        :return:
        """
        # assume user wants to continue existing the task, but reset the timer
        # so this does not clear the selection in task_list_menu or current_task
        # if the timer is not stopped i.e., is_timer_running is running/paused
        if self.is_timer_running != TimerStatus.STOPPED:
            # change the running status
            self.is_timer_running = TimerStatus.STOPPED
            self.task_list_menu.configure(state="normal")
            self.task_start_time = None
            self.segment_start = None
            self.task_end_time = None
            self.segment_end = None
            self.timer_text.set("00:00:00")
            self.notes_textbox.delete("1.0", "end") # clear notes
            self._show_placeholder() # show placeholder text
            self.start_btn.configure(text="▶")
            # as we are updating self.is_timer_running = TimerStatus.STOPPED, _update_timer() will stop as it runs only when the timer is running
            # self._update_timer()
            # to remove focus from notes entry field if the notes were being typed
            self.app.focus()
            if self.timer_running_queue:
                self.app.after_cancel(self.timer_running_queue)
            # we use reset_timer method inside the end_timer method too to avoid code repetition as there are many common operations between both the methods,
            # however, the status for both the methods is diff
            # for reset_timer -> "Reset"
            # for end_timer -> "Saved"/"Error"
            # to address different status for both methods we have a check before updating status
            # we pass status "Reset" for reset_btn call
            # we pass nothing for call from end_timer
            if status:
                self._update_status_label(status, 0)
        # print("Timer reset")


    def _hide_app_window(self):
        """
        hides the app window
        used for custom_minimize button
        """
        self.app.withdraw()


    def _start_drag(self, event):
        """
        gets the initial absolute coordinates of the app window, mouse pointer wrt to the screen
        """
        # Store the initial ABSOLUTE screen coordinates of the mouse click
        self.start_mouse_x_root = event.x_root
        self.start_mouse_y_root = event.y_root

        # Store the initial ABSOLUTE screen coordinates of the app window's top-left corner
        self.start_window_x_root = self.app.winfo_x()
        self.start_window_y_root = self.app.winfo_y()


    def _do_drag(self, event):
        """
        handles repositioning of the app window on mouse hold and drag
        """
        # calculate the total displacement (change) of the mouse from its starting point
        deltax_root = event.x_root - self.start_mouse_x_root
        deltay_root = event.y_root - self.start_mouse_y_root

        # calculate the new absolute window position
        window_new_x = self.start_window_x_root + deltax_root
        window_new_y = self.start_window_y_root + deltay_root

        # reposition the app window at the new coordinates
        self.app.geometry(f"+{window_new_x}+{window_new_y}")


    def _get_resource_path(self, file_name):
        """
        Retrieve the absolute path to resource (file_name) for dev and for PyInstaller (.exe)
        This only returns the absolute file path and does not check 'if the file actually exists/valid/corrupted/readable', just path construction
        :param file_name: str file name of the asset whose path is to be retrieved
        :return: file path of the asset
        """
        if hasattr(sys, "_MEIPASS"):
            # app is running as a PyInstaller bundle if the sys._MEIPASS (Multi-Executable Installer) exists
            base_path = sys._MEIPASS
        else:
            # app is running as a script
            # base_path is the current directory where the script is running from/located in
            # asset should also be in the same directory
            base_path = os.path.abspath(".")
        # print(f"Resolved {file_name} path: {os.path.join(base_path, file_name)}")
        return os.path.join(base_path, file_name)


    def _get_icon(self, icon_name) -> ImageFile:
        """
        checks the existence of icon at the path returned by _get_resource_path() method
        and if the icon file is valid, readable, not corrupted
        if it exists and valid, returns the icon
        else returns a fallback icon created with ImageDraw
        :param str icon_name
        :return: icon_image
        """
        # get the app icon path
        app_icon_path = self._get_resource_path(icon_name)
        # print(f"{app_icon_path=}")

        if os.path.exists(app_icon_path):
            icon_image = Image.open(app_icon_path)
        else:
            # if any error with app_icon, return a blank image with app initials

            # create a blank image with a blue background
            icon_image = Image.new(mode="RGB", size=(36, 36), color="#05428b")
            # get the initials of the app name to be written into blank image
            # default in split is by " " space
            app_name_initials = "".join([word[0] for word in self.app_title.split()])
            # create an image drawer object to write app name 'TK' for timekeeper to the blank image
            drawer = ImageDraw.Draw(icon_image)
            # get a font to draw the app initials into blank image
            # drawer.getfont() -> this gives 'self._draw(no_color_updates=True) # faster drawing without color changes'
            # image_font = ImageDraw.Draw(Image.new("RGB", (1, 1))).getfont()
            # draw app initials on to the blank image
            drawer.text((10, 10), text=app_name_initials, fill="white")

        return icon_image

    #---------system tray icon [start]---------

    def _initialize_systray_icon(self):
        """
        Initializes a python sys tray (pystray) icon
        :return:
        """

        # 1. get the path of the sys tray icon image file -> _get_resource_path()
        # 2. check if the image file exists and is valid at the path -> _get_systray_icon()
        # 3. create menu items -> pystray.MenuItem()
        # 4. create sys tray icon -> pystray.Icon()
        # 5. run the sys tray icon in a loop on a separate thread -> self.systray_thread = threading.Thread()
        try:
            # create menu items for the sys tray icon right-click
            # default True to make it the default action on single click with LMB on the sys tray icon
            menu_items = (
                pystray.MenuItem(f"Open {self.app_title}", self._show_app_window, default=True),
                pystray.MenuItem("Hide", self._hide_app_window),
                pystray.MenuItem("Quit", self._quit_app)
            )

            # get the image file to use as icon
            icon_image = self._get_icon(self.app_icon)
            # create the systray icon with pystray
            # name="time_keeper_widget" -> used by the os/pyinstaller
            # icon_image -> icon shown in sys tray
            # f"{self.app_title} Widget" -> title/tooltip that shows when mouse is hovered on the icon
            self.systray_icon = pystray.Icon("time_keeper_widget", icon_image, f"{self.app_title} Widget", menu=menu_items)
            # start the sys tray icon loop
            self.systray_icon.run_detached()
        except Exception as e:
            print(f"FATAL ERROR: System tray icon creation/run failed: {e}")
            self.app.destroy()


    def _show_app_window(self):
        """
        Opens the app window and brings to the front
        :return:
        """

        # check if the app window is visible or not
        is_app_visible = self.app.winfo_ismapped()

        if not is_app_visible:
            self.app.deiconify()

        self.app.after(0, lambda: self.app.attributes('-topmost', True))
        self.app.after(10, lambda: self.app.attributes('-topmost', False))  # Release after 500ms (150+500)


    def _quit_app(self):
        """
        Quit the app entirely
        """

        # stop the timer
        self._end_timer()

        if self.systray_icon:
            self.systray_icon.stop()
            self.systray_icon = None

        # destroy ctk window and exit the mainloop
        self.app.destroy()

    # ---------system tray icon [end]---------

    def _open_excel_file(self):
        """
        Opens the Excel file using the default system application
        Provides user feedback via the status label
        """
        if not os.path.exists(self.excel_file):
            # if the file does not exist
            self._update_status_label("Error", 1)
            return

        try:
            if sys.platform.startswith('win'):
                # Windows: uses the default application for the file type
                os.startfile(self.excel_file)
                self._update_status_label("Open", 0)
            elif sys.platform.startswith('darwin'):
                # macOS: uses the 'open' command
                import subprocess
                subprocess.run(['open', self.excel_file], check=True)
                self._update_status_label("Opened", 0)
            elif sys.platform.startswith('linux'):
                # Linux: uses 'xdg-open' which opens with the default app
                import subprocess
                subprocess.run(['xdg-open', self.excel_file], check=True)
                self._update_status_label("Opened", 0)
            else:
                self._update_status_label("Error", 1)

        except (FileNotFoundError, Exception) as e:
            # This might happen if the command itself (e.g., 'open', 'xdg-open') is not found
            self._update_status_label("Error", 1)
            print(f"Error opening the file: {e}")


    def _build_ui(self):
        """
        To build the widgets of the app
        """
        # toolbar to show app name and minimize button
        toolbar_frame = ctk.CTkFrame(self.app, height=30, fg_color="#2c2c2c", corner_radius=0)
        toolbar_frame.grid(row=1, column=1, columnspan=3, sticky="we")

        # app logo for toolbar
        # logo_img = ctk.CTkImage(light_image=Image.open("14.ico"), dark_image=Image.open("14.ico"))
        # logo_label = ctk.CTkLabel(toolbar_frame, image=logo_img, text=None)
        # logo_label.grid(row=1, column=1, padx=(10,5), pady=5)

        title_label = ctk.CTkLabel(toolbar_frame, text=self.app_title, font=ctk.CTkFont(size=12))
        title_label.grid(row=1, column=2, sticky="w", pady=5, padx=10)

        custom_minimize_btn = ctk.CTkButton(toolbar_frame, text="\u2013", fg_color="#343638",
                                            hover_color="#585a5c", width=40, height=20,
                                            cursor="hand2", font=("Segoe UI Symbol", 15),
                                            command=self._hide_app_window)
        custom_minimize_btn.grid(row=1, column=2, sticky="e", padx=5, pady=5)
        # to ensure the custom_minimize button sticks to the right edge
        toolbar_frame.grid_columnconfigure(2, weight=1)

        # binding the mouse events to enable dragging functionality to the toolbar_frame and title_label
        toolbar_frame.bind("<Button-1>", self._start_drag)
        toolbar_frame.bind("<B1-Motion>", self._do_drag)
        
        title_label.bind("<Button-1>", self._start_drag)
        title_label.bind("<B1-Motion>", self._do_drag)

        # dropdown menu to choose the tasks from task_list
        self.task_list_menu = ctk.CTkComboBox(self.app, values=self.task_list, command=self._list_menu_callback)
        self.task_list_menu.grid(row=2, column=1, padx=10, pady=(10, 3), sticky="we", columnspan=3)
        # remove the default option displayed from combobox dropdown (defaults to the first option)
        self.task_list_menu.set("")
        # to add a new task on press of the Enter key
        self.task_list_menu.bind("<Return>", self._add_task_on_enter)

        # hint text to show how to add a new task to the task_list
        hint_label = ctk.CTkLabel(self.app, text="Type new task & press Enter", font=("Segoe UI", 12, "bold"), height=5, text_color="#7a848d")
        hint_label.grid(row=3, column=1, columnspan=3, padx=10, sticky="w")

        # status text to show a message on task addition
        self.status_label = ctk.CTkLabel(self.app, text="", font=("Segoe UI", 12, "bold"), height=5)
        self.status_label.grid(row=3, column=3, sticky="e", padx=(0, 12))

        # entry widget to display the running timer
        # set initial text
        initial_timer_text= "00:00:00"
        self.timer_text.set(initial_timer_text)
        self.timer_display = ctk.CTkEntry(self.app, textvariable=self.timer_text, height=75,
                                     font=("Segoe UI Symbol", 40, "bold"), justify="center", state="disabled",
                                     text_color="#9e9e9e")
        self.timer_display.grid(row=4, column=1, columnspan=3, padx=10, pady=10, sticky="we")

        # field to enter notes for the task
        self.notes_textbox = ctk.CTkTextbox(self.app, text_color="#f2f2f2", width=220, height=65,
                                          font=("Segoe UI", 14), wrap="word",
                                          border_width=1, border_color="#4c5154")
        self.notes_textbox.grid(row=5, column=1, columnspan=3, sticky="we", padx=10)

        # show placeholder text at the start
        self._show_placeholder()
        # bind focus-in and focus-out events to handle show/hide of the placeholder text
        self.notes_textbox.bind("<FocusIn>", self._notes_focus_in)
        self.notes_textbox.bind("<FocusOut>", self._notes_focus_out)

        # button frame to hold the buttons and adjust their spacing and widths
        # we have a separate frame for buttons as we have to place 4 buttons in 3 columns
        buttons_frame = ctk.CTkFrame(self.app, fg_color="transparent", height=30)
        buttons_frame.grid(row=6, column=1, columnspan=3, sticky="we")

        buttons_frame.grid_columnconfigure(3, weight=2)
        buttons_frame.grid_columnconfigure(1, weight=1)
        buttons_frame.grid_columnconfigure(2, weight=1)

        # buttons to control the functionality
        # self.start_btn is an instance variable as the text changes ▶ -> ⏸ in run_timer method
        self.start_btn = ctk.CTkButton(buttons_frame, text="▶", command=self._run_timer, cursor="hand2", width=50,
                                       font=("Segoe UI Symbol", 16, "bold"), fg_color="#085bbe",
                                       hover_color="#05428b")
        self.start_btn.grid(padx=10, row=6, column=1, sticky="we", pady=10)

        end_btn = ctk.CTkButton(buttons_frame, text="⏹", width=50, cursor="hand2",
                                font=("Segoe UI Symbol", 16, "bold"), fg_color="#085bbe",
                                command=self._end_timer, hover_color="#05428b")
        end_btn.grid(row=6, column=2, sticky="we")

        reset_btn = ctk.CTkButton(buttons_frame, text="Reset", cursor="hand2", width=60, fg_color="#242424",
                                  border_color="#414449", border_width=1, command=lambda: self._reset_timer("Reset"),
                                  hover_color="#414449")
        reset_btn.grid(padx=(10,5), row=6, column=3, sticky="we")

        excel_btn_icon = ctk.CTkImage(light_image=self._get_icon(self.excel_btn_icon),
                                      dark_image=self._get_icon(self.excel_btn_icon), size=(19,19))
        open_excel_btn = ctk.CTkButton(buttons_frame, image=excel_btn_icon, cursor="hand2", fg_color="#242424",
                                       border_color="#414449", border_width=0, hover_color="#414449", width=1,
                                       text="", command=self._open_excel_file)
        open_excel_btn.grid(row=6, column=4, padx=(0,10), sticky="w", pady=10)

        signature_label = ctk.CTkLabel(self.app, text="akshay;)", text_color="#303030", fg_color="transparent",
                                         font=ctk.CTkFont(size=8, weight="bold", slant="italic"), height=5)
        signature_label.grid(row=6, column=1, sticky="se", columnspan=3)


    def position_window(self):
        # -----------positioning window in the bottom right corner of the screen-----------
        self.app.update_idletasks()

        # Logical vs. Physical Pixels: Screen operates at a physical pixel resolution of 1920x1080, but with 150% DPI scaling, Windows presents a "logical" resolution 1280x720 to applications
        # The coordinates passed to geometry() for positioning might be interpreted as physical pixels by the OS, even if Tkinter is internally working with logical pixels.
        # Solution: By taking the logical positions, multiplying them by dpi_scale_factor (1.5), and then passing those x_pos_physical and y_pos_physical values to app.geometry(), we ensure the window is placed at the correct physical coordinates - bottom right corner

        # position the app window in the bottom right corner of the screen with a margin
        right_margin = 15
        # separate bottom margin to offset the taskbar height
        bottom_margin = 80
        # by trial and error, found app wxh = 240x280 as app.winfo_width() and app.winfo_height() were not giving correct dimensions, probably due to DPI scaling
        app_width = 240
        app_height = 280

        dpi_scale_factor = 1.5

        screen_width_logical = self.app.winfo_screenwidth()
        screen_height_logical = self.app.winfo_screenheight()

        # calculate the logical x y coordinates
        x_logical = screen_width_logical - app_width - right_margin
        y_logical = screen_height_logical - app_height - bottom_margin

        # convert logical coordinates to physical for .geometry()
        x_physical = int(x_logical * dpi_scale_factor)
        y_physical = int(y_logical * dpi_scale_factor)

        # set the app window's position
        self.app.geometry(f"+{x_physical}+{y_physical}")


app = TaskTimer()
