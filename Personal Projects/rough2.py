import customtkinter as ctk
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import datetime as dt
from enum import Enum
import pystray
from PIL import Image as PILImage
import threading


# --- ENUM ---
class TimerStatus(Enum):
    RUNNING = 1
    PAUSED = 2
    STOPPED = 3


# --- MAIN APP CLASS ---
class TaskTimer:
    def __init__(self) -> None:
        # Set appearance mode
        ctk.set_appearance_mode("dark")
        self.app = ctk.CTk()
        self.app.title("Task Timer")
        self.app.resizable(False, False)

        # Excel setup
        self.excel_file = "Task_Timer.xlsx"
        self.excel_tasks_sheet = "Tasks"
        self.tasks_col_name = "Tasks"
        self.excel_time_sheet = "Time"

        # UI variables
        self.task_list = self._get_task_list()
        self.current_task = ""
        self.timer_text = ctk.StringVar()
        self.is_timer_running = TimerStatus.STOPPED
        self.seconds_elapsed = 0
        self.start_time = None
        self.end_time = None

        # UI elements
        self.task_list_menu = None
        self.new_task_status_label = None
        self.timer_display = None
        self.start_btn = None
        self.notes_entry = None

        # Status update queue
        self.status_update_queue = None

        # Build UI
        self.build_ui()

        # Position window
        self.position_window()

        # Setup system tray icon in background thread
        self.tray_icon = None
        self.create_tray_icon()
        threading.Thread(target=self.tray_icon.run, daemon=True).start()

        # Override close button to hide instead of quit
        self.app.protocol("WM_DELETE_WINDOW", self.hide_window)

        # Start main loop
        self.app.mainloop()

    # --- SYSTEM TRAY METHODS ---
    def create_tray_icon(self):
        def on_restore(icon, item):
            self.show_window()

        def on_exit(icon, item):
            self.app.destroy()  # Stop Tkinter
            icon.stop()         # Stop tray

        # Create a simple blue square icon
        icon_image = PILImage.new('RGB', (64, 64), color=(0, 120, 215))

        # Menu
        menu = (
            pystray.MenuItem("Open", on_restore),
            pystray.MenuItem("Exit", on_exit)
        )

        self.tray_icon = pystray.Icon("TaskTimer", icon_image, "Task Timer", menu)

    def show_window(self):
        """Restore window from tray"""
        self.app.deiconify()
        self.app.lift()
        self.app.focus_force()
        self.update_tooltip()  # Refresh tooltip after opening

    def hide_window(self):
        """Hide window but keep running in background"""
        self.app.withdraw()
        self.update_tooltip()  # Update tooltip if timer is running

    def update_tooltip(self):
        """Update tooltip dynamically based on timer status"""
        if self.is_timer_running == TimerStatus.RUNNING:
            elapsed = dt.datetime.now() - self.start_time
            hours, remainder = divmod(int(elapsed.total_seconds()), 3600)
            minutes, seconds = divmod(remainder, 60)
            self.tray_icon.tooltip = f"{self.current_task} | {hours:02}:{minutes:02}:{seconds:02}"
        else:
            self.tray_icon.tooltip = "Task Timer - Idle"

    # --- TIMER LOGIC ---
    def _update_timer(self):
        if self.is_timer_running == TimerStatus.RUNNING:
            self.seconds_elapsed += 1
            hours, remainder = divmod(self.seconds_elapsed, 3600)
            minutes, seconds = divmod(remainder, 60)
            self.timer_text.set(f"{hours:02}:{minutes:02}:{seconds:02}")
            self.update_tooltip()
            self.app.after(1000, self._update_timer)

    def run_timer(self):
        if self.current_task:
            if self.is_timer_running != TimerStatus.RUNNING:
                self.is_timer_running = TimerStatus.RUNNING
                self.start_btn.configure(text="⏸")
                self.task_list_menu.configure(state="disabled")
                self.start_time = dt.datetime.now()
                self._update_timer()
            else:
                self.is_timer_running = TimerStatus.PAUSED
                self.start_btn.configure(text="▶")
        else:
            self._update_task_status_label("Select", 1)
        self.app.focus()

    def end_timer(self):
        if self.is_timer_running != TimerStatus.STOPPED:
            self.end_time = dt.datetime.now()
            self._append_data_to_excel(
                self.excel_time_sheet,
                Date=f"{self.start_time:%d-%b-%Y}",
                Task=self.current_task,
                Duration=self._humanize_time(),
                Notes=self.notes_entry.get("1.0", "end-1c"),
                Start_Time=f"{self.start_time:%I:%M %p}",
                End_Time=f"{self.end_time:%I:%M %p}",
                Total_Seconds=self.seconds_elapsed
            )
            self.reset_timer()
        print("Timer stopped")

    def reset_timer(self):
        if self.is_timer_running != TimerStatus.STOPPED:
            self.is_timer_running = TimerStatus.STOPPED
            self.seconds_elapsed = 0
            self.start_btn.configure(text="▶")
            self.timer_text.set("00:00:00")
            self.task_list_menu.configure(state="normal")
            self.app.focus()
        self.update_tooltip()
        print("Timer reset")

    def _humanize_time(self):
        difference = self.end_time - self.start_time
        diff_seconds = difference.total_seconds()
        hours, remainder = divmod(diff_seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        parts = []
        if hours:
            parts.append(f"{int(hours)}h")
        if minutes:
            parts.append(f"{int(minutes)}m")
        return ":".join(parts)

    # --- EXCEL FUNCTIONS ---
    def _get_task_list(self):
        tasks_list = []
        if os.path.exists(self.excel_file) and os.path.getsize(self.excel_file) > 0:
            try:
                tasks_df = pd.read_excel(self.excel_file, sheet_name=self.excel_tasks_sheet)
                if not tasks_df.empty:
                    active_tasks = tasks_df[
                        tasks_df[self.tasks_col_name].notna() &
                        (tasks_df["Status"].str.lower() == "active")
                    ]
                    tasks_list = active_tasks[self.tasks_col_name].to_list()
            except Exception as e:
                print(f"Error reading task list: {e}")
        return [" <Add new task...>"] + tasks_list

    def _append_data_to_excel(self, sheet_name, **kwargs):
        try:
            if os.path.exists(self.excel_file):
                try:
                    wb = load_workbook(self.excel_file)
                except Exception:
                    wb = Workbook()
                    for s in wb.sheetnames:
                        wb.remove(wb[s])
            else:
                wb = Workbook()
                for s in wb.sheetnames:
                    wb.remove(wb[s])

            if sheet_name not in wb.sheetnames:
                sheet = wb.create_sheet(sheet_name)
                sheet.append(list(kwargs.keys()))
            else:
                sheet = wb[sheet_name]

            sheet.append(list(kwargs.values()))
            wb.save(self.excel_file)
            wb.close()
            return True
        except Exception as e:
            print(f"Error appending data: {e}")
            return False

    # --- TASK INPUT HANDLING ---
    def list_menu_callback(self, choice):
        if choice == " <Add new task...>":
            self.task_list_menu.set("")
            self.task_list_menu.focus()
        else:
            self.current_task = choice
            self.app.focus()

    def _update_task_status_label(self, status: str, code: int):
        if self.status_update_queue:
            self.app.after_cancel(self.status_update_queue)
        if code == 1:
            self.new_task_status_label.configure(text=status, text_color="#b54747")
        else:
            self.new_task_status_label.configure(text=status, text_color="#009933")
        if status:
            self.status_update_queue = self.app.after(3000, lambda: self._update_task_status_label("", 1))

    def _add_task_on_enter(self, event):
        new_task = self.task_list_menu.get().strip()
        if new_task:
            new_task = new_task[0].upper() + new_task[1:]
            if new_task not in self.task_list:
                now = dt.datetime.now()
                write_status = self._append_data_to_excel(
                    self.excel_tasks_sheet,
                    Tasks=new_task,
                    Status="Active",
                    Added_On=f"{now:%b %d, %Y T%I:%M %p}"
                )
                if write_status:
                    self.task_list = self._get_task_list()
                    self.task_list.sort()
                    self.task_list_menu.configure(values=self.task_list)
                    self.task_list_menu.set(new_task)
                    self.current_task = new_task
                    self._update_task_status_label("Added", 0)
                    self.app.focus()
                else:
                    self._update_task_status_label("Error!", 1)
            else:
                self._update_task_status_label("Exists!", 1)
        else:
            self._update_task_status_label("Empty!", 1)

    # --- UI BUILDING ---
    def build_ui(self):
        # -- Dropdown --
        self.task_list_menu = ctk.CTkComboBox(self.app, values=self.task_list, command=self.list_menu_callback)
        self.task_list_menu.grid(row=1, column=1, padx=10, pady=(10, 3), sticky="we", columnspan=3)
        self.task_list_menu.set("")
        self.task_list_menu.bind("<Return>", self._add_task_on_enter)

        # -- Hint Label --
        hint_label = ctk.CTkLabel(self.app, text="Type new task & press Enter", font=("Segoe UI", 12, "bold"), height=5,
                                  text_color="#7a848d")
        hint_label.grid(row=2, column=1, columnspan=3, padx=10, sticky="w")

        # -- Status Label --
        self.new_task_status_label = ctk.CTkLabel(self.app, text="", text_color="#009933",
                                                  font=("Segoe UI", 12, "bold"), height=5)
        self.new_task_status_label.grid(row=2, column=3, sticky="e", padx=(0, 12))

        # -- Timer Display --
        initial_timer_text = "00:00:00"
        self.timer_text.set(initial_timer_text)
        self.timer_display = ctk.CTkEntry(self.app, textvariable=self.timer_text, height=75,
                                          font=("Segoe UI Symbol", 40, "bold"), justify="center", state="disabled",
                                          text_color="#9e9e9e")
        self.timer_display.grid(row=3, column=1, columnspan=3, padx=10, pady=10, sticky="we")

        # -- Notes Box --
        self.notes_entry = ctk.CTkTextbox(self.app, text_color="#f2f2f2", width=220, height=65,
                                          font=("Segoe UI", 14), border_width=1, border_color="#4c5154")
        self.notes_entry.grid(row=4, column=1, columnspan=3, sticky="we", padx=10)

        # -- Buttons --
        self.start_btn = ctk.CTkButton(self.app, text="▶", command=self.run_timer, cursor="hand2", width=40,
                                       font=("Segoe UI Symbol", 16, "bold"), fg_color="#085bbe")
        self.start_btn.grid(padx=10, row=5, column=1, sticky="we", pady=10)

        end_btn = ctk.CTkButton(self.app, text="⏹", width=40, cursor="hand2", font=("Segoe UI Symbol", 16, "bold"),
                                fg_color="#085bbe", command=self.end_timer)
        end_btn.grid(row=5, column=2, sticky="we")

        reset_btn = ctk.CTkButton(self.app, text="Reset", cursor="hand2", width=60, fg_color="#242424",
                                  border_color="#414449", border_width=1, command=self.reset_timer)
        reset_btn.grid(padx=10, row=5, column=3, sticky="we")

    def position_window(self):
        self.app.update_idletasks()
        dpi_scale_factor = 1.5
        app_width, app_height = 240, 280
        screen_width = self.app.winfo_screenwidth()
        screen_height = self.app.winfo_screenheight()
        x_logical = screen_width - app_width - 20
        y_logical = screen_height - app_height - 80
        x_physical = int(x_logical * dpi_scale_factor)
        y_physical = int(y_logical * dpi_scale_factor)
        self.app.geometry(f"+{x_physical}+{y_physical}")


# --- RUN APP ---
if __name__ == "__main__":
    app = TaskTimer()